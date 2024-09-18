#Doğrulama Kodu
import requests
from bs4 import BeautifulSoup
url = "https://docs.google.com/spreadsheets/d/1AP9EFAOthh5gsHjBCDHoUMhpef4MSxYg6wBN0ndTcnA/edit#gid=0"
response = requests.get(url)
html_content = response.content
soup = BeautifulSoup(html_content, "html.parser")
first_cell = soup.find("td", {"class": "s2"}).text.strip()
if first_cell != "Aktif":
    exit()
first_cell = soup.find("td", {"class": "s1"}).text.strip()
print(first_cell)




import requests
import pandas as pd
import os
from openpyxl.worksheet.table import Table, TableStyleInfo
from copy import copy
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter



def download_and_merge_excel(url1, url2):
    # İlk Excel dosyasını indir
    response1 = requests.get(url1)
    with open('excel1.xlsx', 'wb') as f1:
        f1.write(response1.content)

    # İkinci Excel dosyasını indir
    response2 = requests.get(url2)
    with open('excel2.xlsx', 'wb') as f2:
        f2.write(response2.content)

    # İki Excel dosyasını birleştir
    df1 = pd.read_excel('excel1.xlsx')
    df2 = pd.read_excel('excel2.xlsx')

    merged_df = pd.concat([df1, df2], ignore_index=True)

    # Birleştirilmiş DataFrame'i yeni bir Excel dosyasına yaz
    merged_df.to_excel('UrunListesi.xlsx', index=False)

    # İndirilen Excel dosyalarını sil
    
    os.remove('excel1.xlsx')
    os.remove('excel2.xlsx')

if __name__ == "__main__":
    url1 = "https://task.haydigiy.com/FaprikaXls/ZZ5GMS/1/"
    url2 = "https://task.haydigiy.com/FaprikaXls/ZZ5GMS/2/"

    download_and_merge_excel(url1, url2)







import pandas as pd

# Birleştirilmiş Excel dosyasını oku
df_merged = pd.read_excel('UrunListesi.xlsx')

# İstenmeyen sütunları belirle
columns_to_keep = ["ModelKodu", "UrunAdi", "StokAdedi", "AmazonKodu", "Aciklama"]

# İstenmeyen sütunları seç
df_merged = df_merged[columns_to_keep]

# 'Aciklama' sütunundaki dolu olan satırları filtrele
df_merged = df_merged[df_merged['Aciklama'].isna()]

# Güncellenmiş DataFrame'i aynı Excel dosyasının üstüne yaz
df_merged.to_excel('UrunListesi.xlsx', index=False)













# "ModelKodu_Copy" Excel dosyasını oku
df_copy = pd.read_excel('UrunListesi.xlsx')

# "ModelKodu_Copy_Unique" sütununu sil
df_copy = df_copy.drop(['Aciklama'], axis=1, errors='ignore')

# Güncellenmiş DataFrame'i "ModelKodu_Copy" Excel dosyasına yaz
df_copy.to_excel('UrunListesi.xlsx', index=False)






# Veriyi Okuma
df = pd.read_excel('UrunListesi.xlsx')

# "UrunAdi" Sütunundaki "StokAdedi" Değerlerinin Toplamını Hesapla
df['StokAdedi2'] = df.groupby('UrunAdi')['StokAdedi'].transform('sum')


# "StokAdedi" Sütununu Sil
df = df.drop(['StokAdedi'], axis=1, errors='ignore')

# Yenilenen değerleri teke düşür
df = df.drop_duplicates()

# Sonucu Mevcut Excel Dosyasının Üzerine Kaydetme
df.to_excel('UrunListesi.xlsx', index=False)






# Veriyi Okuma
df = pd.read_excel('UrunListesi.xlsx')

# "ModelKodu" sütununu kopyala
model_kodu_copy = df['ModelKodu'].copy()

# Yeni bir DataFrame oluştur ve "ModelKodu" sütununu ekleyin
new_df = pd.DataFrame({'ModelKodu_Copy': model_kodu_copy})


# Yeni DataFrame'i yeni bir Excel dosyasına kaydet
new_df.to_excel('ModelKodu_Copy.xlsx', index=False)




# Yeni Excel dosyasını oku
df_copy = pd.read_excel('ModelKodu_Copy.xlsx')

# "ModelKodu_Copy" sütunundaki benzersiz değerleri al
unique_values = df_copy['ModelKodu_Copy'].unique()

# Yeni bir DataFrame oluştur ve benzersiz değerleri içerecek şekilde "ModelKodu_Copy" sütununu ekleyin
unique_df = pd.DataFrame({'ModelKodu_Copy_Unique': unique_values})

# Yeni DataFrame'i yeni bir Excel dosyasına kaydet
unique_df.to_excel('ModelKodu_Copy.xlsx', index=False)






# "ModelKodu_Copy" ve "ModelKodu_Copy_Unique" Excel dosyalarını oku
df_copy = pd.read_excel('ModelKodu_Copy.xlsx')
df_urun = pd.read_excel('UrunListesi.xlsx')

# "ModelKodu_Copy_Unique" sütunundaki değerleri döngüye al
for index, row in df_copy.iterrows():
    model_kodu_unique = row['ModelKodu_Copy_Unique']

    # "UrunListesi" dosyasında "ModelKodu" sütununda arama yap
    matching_row = df_urun.loc[df_urun['ModelKodu'] == model_kodu_unique]

    # Eğer karşılık gelen bir değer bulunduysa, ilgili sütunları "ModelKodu_Copy" dosyasına yaz
    if not matching_row.empty:
        df_copy.at[index, 'Ürün Adı'] = matching_row['UrunAdi'].values[0]
        df_copy.at[index, 'Stok Adedi'] = matching_row['StokAdedi2'].values[0]
        df_copy.at[index, 'Raf Kodu'] = matching_row['AmazonKodu'].values[0]

# Güncellenmiş DataFrame'i "ModelKodu_Copy" Excel dosyasına yaz
df_copy.to_excel('ModelKodu_Copy.xlsx', index=False)





# "ModelKodu_Copy" Excel dosyasını oku
df_copy = pd.read_excel('ModelKodu_Copy.xlsx')

# "ModelKodu_Copy_Unique" sütununu sil
df_copy = df_copy.drop(['ModelKodu_Copy_Unique'], axis=1, errors='ignore')

# Güncellenmiş DataFrame'i "ModelKodu_Copy" Excel dosyasına yaz
df_copy.to_excel('ModelKodu_Copy.xlsx', index=False)







# "ModelKodu_Copy" Excel dosyasını oku
df_copy = pd.read_excel('ModelKodu_Copy.xlsx')

# "RafKodu" sütununu kopyala ve hemen yanına yapıştır
df_copy['RafKodu_Copy'] = df_copy['Raf Kodu']

# "-" karakterinden sonrasını temizle
df_copy['RafKodu_Copy'] = df_copy['RafKodu_Copy'].apply(lambda x: x.split('-')[0] if isinstance(x, str) else x)

# Tüm verileri sayıya dönüştür
df_copy['RafKodu_Copy'] = pd.to_numeric(df_copy['RafKodu_Copy'], errors='coerce')

# Tüm tabloyu "RafKodu_Copy" sütununa göre küçükten büyüğe sırala
df_copy = df_copy.sort_values(by='RafKodu_Copy', ascending=True)

# Güncellenmiş DataFrame'i "ModelKodu_Copy" Excel dosyasına yaz
df_copy.to_excel('ModelKodu_Copy.xlsx', index=False)



# "ModelKodu_Copy" Excel dosyasını oku
df_copy = pd.read_excel('ModelKodu_Copy.xlsx')

# "ModelKodu_Copy_Unique" sütununu sil
df_copy = df_copy.drop(['RafKodu_Copy'], axis=1, errors='ignore')

# Güncellenmiş DataFrame'i "ModelKodu_Copy" Excel dosyasına yaz
df_copy.to_excel('ModelKodu_Copy.xlsx', index=False)











import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side

# "ModelKodu_Copy" Excel dosyasını oku
df_copy = pd.read_excel('ModelKodu_Copy.xlsx')

# "RafKodu" sütununu başa al
column_order = ['Raf Kodu', 'Ürün Adı', 'Stok Adedi']
df_copy = df_copy[column_order]

# Güncellenmiş DataFrame'i "ModelKodu_Copy" Excel dosyasına yaz
df_copy.to_excel('ModelKodu_Copy.xlsx', index=False)











# Sonuç dosyasını yükle
file_path = "ModelKodu_Copy.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# Hücreleri ortala ve ortaya hizala
for row in main_sheet.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")


# Değişiklikleri kaydet
wb.save("ModelKodu_Copy.xlsx")







    
    

# Sonuç dosyasını yükle
file_path = "ModelKodu_Copy.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# Kenarlık stili oluştur
border_style = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Hücreleri kalın yap, yazı fontunu 14 yap ve kenarlık ekle
for row in main_sheet.iter_rows():
    for cell in row:
        cell.font = Font(bold=True, size=14)
        cell.border = border_style

# Değişiklikleri kaydet
wb.save("ModelKodu_Copy.xlsx")



    

# Sonuç dosyasını yükle
file_path = "ModelKodu_Copy.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# "RAF KODU" sütununu 45 piksel yap
main_sheet.column_dimensions["C"].width = 45

# Tüm hücreleri en uygun sütun genişliği olarak ayarla
for column in main_sheet.columns:
    max_length = 0
    column = list(column)
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    main_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

# Değişiklikleri kaydet
wb.save("ModelKodu_Copy.xlsx")




# Sonuç dosyasını yükle
file_path = "ModelKodu_Copy.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# İlk sütunu (A sütunu) 45 piksel genişliğinde yap
main_sheet.column_dimensions["A"].width = 45
main_sheet.column_dimensions["C"].width = 14
main_sheet.column_dimensions["G"].width = 14
main_sheet.column_dimensions["D"].width = 9
main_sheet.column_dimensions["F"].width = 5

# Değişiklikleri kaydet
wb.save("ModelKodu_Copy.xlsx")









    

# Sonuç dosyasını yükle
file_path = "ModelKodu_Copy.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# Tüm hücrelere "Metni Kaydır" formatını uygula
for row in main_sheet.iter_rows():
    for cell in row:
        new_alignment = copy(cell.alignment)
        new_alignment.wrap_text = True
        cell.alignment = new_alignment

# Değişiklikleri kaydet
wb.save("ModelKodu_Copy.xlsx")








    

# Sonuç dosyasını yükle
file_path = "ModelKodu_Copy.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

# Tabloyu oluşturma
table = Table(displayName="MyTable", ref=main_sheet.dimensions)

# Tablo stili oluşturma (gri-beyaz)
style = TableStyleInfo(
    name="TableStyleLight1", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=True
)

# Tabloya stil atama
table.tableStyleInfo = style

# Tabloyu sayfaya ekleme
main_sheet.add_table(table)

# Değişiklikleri kaydetme
wb.save("ModelKodu_Copy.xlsx")





import os

# "UrunListesi" Excel dosyasını sil
if os.path.exists('UrunListesi.xlsx'):
    os.remove('UrunListesi.xlsx')

# "ModelKodu_copy" adlı Excel dosyasının adını "Özelliksiz Ürünler" olarak değiştir
if os.path.exists('ModelKodu_copy.xlsx'):
    os.rename('ModelKodu_copy.xlsx', 'Özelliksiz Ürünler.xlsx')

    



